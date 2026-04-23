"""
Merris KB Text Extractor
========================
Extracts text from all PDF / XLSX / JSON / DOCX files in the M01-M14 KB folders
and saves a .txt file alongside each source file.

The Node.js embed script then reads these .txt files — no pdfjs-dist needed.

Usage:
    python extract-kb-text.py --dir "C:/Users/.../kb" [--modules M01,M02] [--rebuild]

Requirements:
    pip install pymupdf openpyxl python-docx

PDF engine priority:
    1. pymupdf (fitz)   — best CMap/ToUnicode support, fixes article-number misreads
    2. pdfplumber        — fallback if pymupdf not installed

Handles:
    .pdf   → pymupdf (primary) or pdfplumber (fallback)
    .xlsx  → openpyxl
    .csv   → built-in csv module
    .json  → built-in json module
    .docx  → python-docx
    .txt   → copied as-is
"""

import argparse
import csv
import json
import os
import sys
import time
from pathlib import Path

# ============================================================
# Dependency check
# ============================================================

def check_deps():
    missing = []
    has_pymupdf = False
    has_pdfplumber = False

    try:
        import fitz  # noqa  (pymupdf)
        has_pymupdf = True
    except ImportError:
        missing.append("pymupdf")

    try:
        import pdfplumber  # noqa
        has_pdfplumber = True
    except ImportError:
        if not has_pymupdf:
            missing.append("pdfplumber")

    try:
        import openpyxl  # noqa
    except ImportError:
        missing.append("openpyxl")

    if missing:
        print(f"Missing packages: {', '.join(missing)}")
        print(f"Install with:  pip install {' '.join(missing)}")
        sys.exit(1)

    if not has_pymupdf:
        print("WARNING: pymupdf not installed — falling back to pdfplumber.")
        print("         pdfplumber may misread article numbers in some EU regulatory PDFs.")
        print("         Install pymupdf for better accuracy: pip install pymupdf")

    return has_pymupdf

HAS_PYMUPDF = check_deps()

if HAS_PYMUPDF:
    import fitz  # pymupdf

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

import openpyxl

# python-docx is optional
try:
    import docx as python_docx
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

# ============================================================
# Text extractors
# ============================================================

MAX_PAGES = 200  # cap per PDF to keep memory/time reasonable


def extract_pdf_pymupdf(path: Path) -> str:
    """
    Primary PDF extractor — uses MuPDF engine (pymupdf/fitz).
    Best ToUnicode CMap support: correctly handles article numbers and
    special characters in EU regulatory documents.
    """
    parts = []
    try:
        with fitz.open(str(path)) as doc:
            total = doc.page_count
            pages_to_read = min(total, MAX_PAGES)
            for page_num in range(pages_to_read):
                page = doc[page_num]
                # get_text("text") preserves layout; blocks mode is better for tables
                text = page.get_text("text")
                if text and text.strip():
                    parts.append(text.strip())
            if total > MAX_PAGES:
                parts.append(f"\n[Note: document has {total} pages; only first {MAX_PAGES} extracted]")
    except Exception as e:
        return f"[PyMuPDF extraction failed: {e}]"
    return "\n\n".join(parts)


def extract_pdf_pdfplumber(path: Path) -> str:
    """
    Fallback PDF extractor — uses pdfplumber/pdfminer.
    Good for tables; weaker CMap support for custom-encoded fonts.
    """
    parts = []
    try:
        with pdfplumber.open(str(path)) as pdf:
            total = len(pdf.pages)
            pages_to_read = min(total, MAX_PAGES)
            for page in pdf.pages[:pages_to_read]:
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        for row in table:
                            if row:
                                parts.append("\t".join(str(cell or "") for cell in row))
                text = page.extract_text(x_tolerance=2, y_tolerance=2)
                if text:
                    parts.append(text)
            if total > MAX_PAGES:
                parts.append(f"\n[Note: document has {total} pages; only first {MAX_PAGES} extracted]")
    except Exception as e:
        return f"[pdfplumber extraction failed: {e}]"
    return "\n".join(parts)


def extract_pdf(path: Path) -> str:
    if HAS_PYMUPDF:
        result = extract_pdf_pymupdf(path)
        # If pymupdf fails, try pdfplumber as fallback
        if result.startswith("[PyMuPDF extraction failed") and HAS_PDFPLUMBER:
            print(f"\n  pymupdf failed, retrying with pdfplumber... ", end="", flush=True)
            result = extract_pdf_pdfplumber(path)
        return result
    elif HAS_PDFPLUMBER:
        return extract_pdf_pdfplumber(path)
    else:
        return "[No PDF extractor available — install pymupdf: pip install pymupdf]"


def extract_xlsx(path: Path) -> str:
    parts = []
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"Sheet: {sheet_name}")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    parts.append("\t".join(cells))
        wb.close()
    except Exception as e:
        return f"[XLSX extraction failed: {e}]"
    return "\n".join(parts)


def extract_csv(path: Path) -> str:
    parts = []
    try:
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.reader(f)
            for row in reader:
                parts.append("\t".join(row))
    except Exception as e:
        return f"[CSV extraction failed: {e}]"
    return "\n".join(parts)


def extract_json(path: Path) -> str:
    try:
        with open(path, encoding="utf-8", errors="replace") as f:
            data = json.load(f)
        if isinstance(data, list):
            lines = []
            for item in data:
                if isinstance(item, dict):
                    lines.append("\n".join(f"{k}: {json.dumps(v, ensure_ascii=False)}" for k, v in item.items()))
                else:
                    lines.append(str(item))
            return "\n\n".join(lines)
        elif isinstance(data, dict):
            return "\n".join(f"{k}: {json.dumps(v, ensure_ascii=False)}" for k, v in data.items())
        else:
            return str(data)
    except Exception as e:
        return f"[JSON extraction failed: {e}]"


def extract_docx(path: Path) -> str:
    if not HAS_DOCX:
        return "[DOCX extraction skipped — install python-docx: pip install python-docx]"
    try:
        doc = python_docx.Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as e:
        return f"[DOCX extraction failed: {e}]"


def extract_text(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return extract_pdf(path)
    elif ext in (".xlsx", ".xls"):
        return extract_xlsx(path)
    elif ext == ".csv":
        return extract_csv(path)
    elif ext == ".json":
        return extract_json(path)
    elif ext == ".docx":
        return extract_docx(path)
    elif ext in (".txt", ".md"):
        try:
            return path.read_text(encoding="utf-8", errors="replace")
        except Exception as e:
            return f"[Read failed: {e}]"
    return ""  # unsupported


# ============================================================
# File discovery
# ============================================================

SUPPORTED = {".pdf", ".xlsx", ".xls", ".csv", ".json", ".docx", ".txt", ".md"}


def discover_files(kb_root: Path, module_filter: list[str]) -> list[Path]:
    files = []
    for entry in sorted(kb_root.iterdir()):
        if not entry.is_dir():
            continue
        if entry.name.startswith("_") or entry.name.startswith("."):
            continue
        if module_filter:
            if not any(entry.name == m or entry.name.startswith(m + "-") or entry.name.startswith(m + "_")
                       for m in module_filter):
                continue
        for fp in sorted(entry.rglob("*")):
            if fp.is_file() and fp.suffix.lower() in SUPPORTED and fp.suffix.lower() != ".txt":
                files.append(fp)
    return files


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Extract text from KB files to .txt sidecar files")
    parser.add_argument("--dir", required=True, help="KB root directory containing M01-M14 folders")
    parser.add_argument("--modules", default="", help="Comma-separated modules e.g. M01,M02 (default: all)")
    parser.add_argument("--rebuild", action="store_true", help="Re-extract even if .txt already exists")
    args = parser.parse_args()

    kb_root = Path(args.dir)
    if not kb_root.is_dir():
        print(f"ERROR: Directory not found: {kb_root}")
        sys.exit(1)

    module_filter = [m.strip() for m in args.modules.split(",") if m.strip()]
    start = time.time()

    print("=" * 60)
    print("Merris KB Text Extractor")
    print(f"PDF engine : {'pymupdf (primary)' if HAS_PYMUPDF else 'pdfplumber (fallback)'}")
    print(f"Source     : {kb_root}")
    print(f"Modules    : {', '.join(module_filter) if module_filter else 'all'}")
    print(f"Rebuild    : {args.rebuild}")
    print("=" * 60)

    files = discover_files(kb_root, module_filter)
    print(f"\nFound {len(files)} files\n")

    success = 0
    skipped = 0
    failed  = 0

    for i, fp in enumerate(files, 1):
        txt_path = fp.with_suffix(".txt")
        rel = fp.relative_to(kb_root)

        if txt_path.exists() and not args.rebuild:
            print(f"[{i}/{len(files)}] SKIP  {rel}  (already extracted)")
            skipped += 1
            continue

        print(f"[{i}/{len(files)}] {rel} ... ", end="", flush=True)
        t0 = time.time()

        try:
            text = extract_text(fp)
            if not text or len(text.strip()) < 20:
                print(f"EMPTY ({time.time()-t0:.1f}s)")
                failed += 1
                continue

            txt_path.write_text(text, encoding="utf-8")
            size_kb = txt_path.stat().st_size // 1024
            print(f"OK  {size_kb}KB  ({time.time()-t0:.1f}s)")
            success += 1
        except Exception as e:
            print(f"FAILED: {e}")
            failed += 1

    elapsed = round(time.time() - start)
    print("\n" + "=" * 60)
    print(f"Done in {elapsed}s")
    print(f"  PDF engine  : {'pymupdf' if HAS_PYMUPDF else 'pdfplumber'}")
    print(f"  Extracted   : {success}")
    print(f"  Skipped     : {skipped}")
    print(f"  Failed      : {failed}")
    print(f"  Total       : {len(files)}")
    print("=" * 60)


if __name__ == "__main__":
    main()
